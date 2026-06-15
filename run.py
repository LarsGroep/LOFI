"""
LOFI Tinder -- main entrypoint.

Full pipeline when using Supabase as data source (GitHub Actions scraper has run):
    1. python run.py --pull            # sync Supabase tinder.artist_cache -> scraper_data/artist_enriched.jsonl
    2. python run.py --seed            # generate profiles for LOFI-booked artists + build centroids + MAB
    3. python run.py --candidates      # generate profiles for candidate artists
    4. python run.py --build-index     # (re)build FAISS index
    5. streamlit run lofi_tinder/app.py

Full pipeline when using local scraper data:
    1. python run.py --enrich          # aggregate all scraper data -> scraper_data/artist_enriched.jsonl
    2-5. same as above

Quick refresh after nightly GitHub Actions run:
    python run.py --pull && python run.py --seed && python run.py --candidates

Other:
    python run.py --stats              # show current state
    python run.py --retrain-mab        # replay all swipes through 14-dim MAB from scratch
    python run.py --export-excel       # export scraper_data to Excel
"""

from __future__ import annotations

import argparse
import json
import sys

# Prevent UnicodeEncodeError on Windows terminals (cp1252) for artist names with diacritics
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

_DATA_DIR        = Path(__file__).parent / "data"
_SCRAPER_DATA    = Path(__file__).parent / "scraper_data"
_PROFILES_FILE   = Path(__file__).parent / "profiles" / "artist_profiles.jsonl"
_CANDIDATES_FILE = _DATA_DIR / "candidates.jsonl"
_ENRICHED_FILE   = _SCRAPER_DATA / "artist_enriched.jsonl"


# ---------------------------------------------------------------------------
# Load enriched data
# ---------------------------------------------------------------------------

def _load_enriched() -> dict[str, dict]:
    if not _ENRICHED_FILE.exists():
        raise FileNotFoundError(
            f"Enriched data not found. Run first: python run.py --enrich"
        )
    records: dict[str, dict] = {}
    for line in _ENRICHED_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            try:
                d = json.loads(line)
                records[d["artist_id"]] = d
            except Exception:
                pass
    return records


def _enriched_to_artist_input(record: dict):
    from lofi_tinder.schemas import ArtistInput
    return ArtistInput(
        artist_id=record["artist_id"],
        name=record["name"],
        enriched=record,
    )


def _save_candidate(artist_input) -> None:
    _DATA_DIR.mkdir(parents=True, exist_ok=True)
    existing = set()
    if _CANDIDATES_FILE.exists():
        for line in _CANDIDATES_FILE.read_text(encoding="utf-8").splitlines():
            if line.strip():
                try:
                    existing.add(json.loads(line)["artist_id"])
                except Exception:
                    pass
    if artist_input.artist_id not in existing:
        with open(_CANDIDATES_FILE, "a", encoding="utf-8") as f:
            f.write(artist_input.model_dump_json() + "\n")


def _flush_profile_embeddings(profiles_map) -> None:
    _PROFILES_FILE.parent.mkdir(parents=True, exist_ok=True)
    existing: dict[str, dict] = {}
    if _PROFILES_FILE.exists():
        for line in _PROFILES_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                try:
                    d = json.loads(line)
                    existing[d["artist_id"]] = d
                except Exception:
                    pass
    for artist_id, profile in profiles_map.items():
        existing[artist_id] = json.loads(profile.model_dump_json())
    with open(_PROFILES_FILE, "w", encoding="utf-8") as f:
        for d in existing.values():
            f.write(json.dumps(d, ensure_ascii=False) + "\n")

    # Push to Supabase so Streamlit Cloud can load profiles without local files
    try:
        from dotenv import load_dotenv
        load_dotenv()
        from lofi_tinder.supabase_client import get_client
        sb = get_client()
        if sb.available:
            pushed = 0
            for artist_id, profile in profiles_map.items():
                sb.save_profile(
                    slug=profile.artist_id,
                    name=profile.name,
                    profile_text=profile.profile_text,
                    embedding=profile.embedding if profile.embedding else None,
                    cosine_dist=profile.cosine_dist_to_centroid,
                )
                pushed += 1
            print(f"  Pushed {pushed} profiles to Supabase tinder.artist_profiles")
        else:
            from lofi_tinder.supabase_client import get_connect_error
            print(f"  Supabase unavailable ({get_connect_error()}) — profiles saved locally only")
    except Exception as e:
        print(f"  Supabase profile push failed: {e}")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel() -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime, timezone

    records = list(_load_enriched().values())
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Artist Overview ----
    ws = wb.active
    ws.title = "Artist Overview"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    cols = [
        ("Name",              lambda r: r["name"]),
        ("LOFI Booked",       lambda r: "YES" if r.get("lofi_booked") else ""),
        ("LOFI Appearances",  lambda r: r.get("lofi_appearance_count") or 0),
        ("Momentum Score",    lambda r: r.get("momentum_score")),
        ("LFM Listeners",     lambda r: (r.get("growth_history") or {}).get("current_listeners")),
        ("LFM Growth % Total",lambda r: (r.get("growth_history") or {}).get("listener_growth_pct_total")),
        ("Days Tracked",      lambda r: (r.get("growth_history") or {}).get("days_tracked")),
        ("PF Fans",           lambda r: r.get("pf_fans")),
        ("Total Bookings",    lambda r: (r.get("booking_stats") or {}).get("total")),
        ("Bookings 12m",      lambda r: (r.get("booking_stats") or {}).get("recent_12m")),
        ("Booking Velocity",  lambda r: (r.get("booking_stats") or {}).get("booking_velocity")),
        ("Countries",         lambda r: (r.get("booking_stats") or {}).get("geo_spread")),
        ("NL Ratio",          lambda r: r.get("nl_ratio")),
        ("Festival Count",    lambda r: (r.get("booking_stats") or {}).get("festival_count")),
        ("Festivals",         lambda r: "; ".join((r.get("festival_history") or [])[:5])),
        ("BP Releases",       lambda r: r.get("beatport_releases")),
        ("BP Labels",         lambda r: "; ".join((r.get("beatport_labels") or [])[:3])),
        ("BP Tier",           lambda r: r.get("beatport_label_tier")),
        ("BP Latest Release", lambda r: r.get("beatport_latest_release")),
        ("Mixcloud Features", lambda r: r.get("mixcloud_appearances")),
        ("Mixcloud Shows",    lambda r: "; ".join((r.get("mixcloud_shows") or [])[:3])),
        ("RA Genre Events",   lambda r: r.get("ra_genre_events")),
        ("RA Genres",         lambda r: "; ".join((r.get("ra_genres") or [])[:4])),
        ("Agency",            lambda r: r.get("agency") or "unknown"),
        ("Agency Tier",       lambda r: r.get("agency_tier") or ""),
        ("Tags",              lambda r: "; ".join((r.get("lastfm_tags") or [])[:5])),
        ("Similar Artists",   lambda r: "; ".join((r.get("lastfm_similar") or [])[:5])),
        ("Spotify URL",       lambda r: r.get("spotify_url") or ""),
        ("Top Cities",        lambda r: "; ".join((r.get("booking_stats") or {}).get("cities", [])[:5])),
    ]

    for col_idx, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(
        sorted(records, key=lambda r: r.get("momentum_score", 0), reverse=True), 2
    ):
        for col_idx, (_, extractor) in enumerate(cols, 1):
            try:
                val = extractor(record)
                if isinstance(val, float):
                    val = round(val, 2)
                ws.cell(row=row_idx, column=col_idx, value=val)
            except Exception:
                pass

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # ---- Sheet 2: Booking History ----
    ws2 = wb.create_sheet("Booking History")
    bh_cols = ["Artist", "Date", "Venue", "City", "Country", "Event Name"]
    for col_idx, h in enumerate(bh_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row_idx = 2
    for record in sorted(records, key=lambda r: r["name"]):
        for ev in (record.get("booking_stats") or {}).get("recent_events") or []:
            ws2.cell(row=row_idx, column=1, value=record["name"])
            ws2.cell(row=row_idx, column=2, value=ev.get("date"))
            ws2.cell(row=row_idx, column=3, value=ev.get("venue"))
            ws2.cell(row=row_idx, column=4, value=ev.get("city"))
            ws2.cell(row=row_idx, column=5, value=ev.get("country"))
            ws2.cell(row=row_idx, column=6, value=ev.get("event_name"))
            row_idx += 1

    # ---- Sheet 3: Growth History ----
    ws3 = wb.create_sheet("Growth History")
    gh_cols = ["Artist", "Snapshot Date", "Listeners", "Playcount"]
    for col_idx, h in enumerate(gh_cols, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row_idx = 2
    for record in sorted(records, key=lambda r: r["name"]):
        for snap in (record.get("growth_history") or {}).get("snapshots") or []:
            ws3.cell(row=row_idx, column=1, value=record["name"])
            ws3.cell(row=row_idx, column=2, value=snap.get("date"))
            ws3.cell(row=row_idx, column=3, value=snap.get("listeners"))
            ws3.cell(row=row_idx, column=4, value=snap.get("playcount"))
            row_idx += 1

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    out_path = _SCRAPER_DATA / f"lofi_artist_data_{ts}.xlsx"
    wb.save(str(out_path))
    print(f"Excel exported: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------


def _cache_row_to_enriched(row: dict) -> dict:
    """Convert a tinder.artist_cache row to the enriched dict format used by run.py."""
    r = {k: v for k, v in row.items() if v is not None}
    # slug is the artist_id in the enriched format
    r["artist_id"] = r.pop("slug", r.get("artist_id", ""))
    # Reconstruct growth_history from flat columns if not present
    if not r.get("growth_history"):
        gh: dict = {}
        if r.get("lastfm_listeners"):
            gh["current_listeners"] = r["lastfm_listeners"]
        if r.get("lastfm_playcount"):
            gh["current_playcount"] = r["lastfm_playcount"]
        if gh:
            r["growth_history"] = gh
    # Remove cache-specific housekeeping fields
    for f in ("cache_updated_at", "last_scraped_at", "artist_id"):
        r.pop(f, None)
    # Restore artist_id (we removed the wrong key above)
    r["artist_id"] = row.get("slug", "")
    return r


def cmd_pull(args) -> None:
    """Pull all artists from Supabase tinder.artist_cache into local artist_enriched.jsonl."""
    from dotenv import load_dotenv
    load_dotenv()
    from lofi_tinder.supabase_client import get_client

    sb = get_client()
    if not sb.available:
        from lofi_tinder.supabase_client import get_connect_error
        print(f"Supabase not available: {get_connect_error()}")
        return

    print("Fetching artists from Supabase tinder.artist_cache...")
    rows = sb.load_artists()
    if not rows:
        print("No artists found in Supabase. Has the GitHub Actions scraper run yet?")
        return

    print(f"  Fetched {len(rows)} artists")

    records: dict[str, dict] = {}
    # Keep existing local records (preserves fields not in Supabase cache)
    if _ENRICHED_FILE.exists():
        for line in _ENRICHED_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                try:
                    d = json.loads(line)
                    records[d["artist_id"]] = d
                except Exception:
                    pass
        print(f"  Merging with {len(records)} existing local records")

    updated = 0
    for row in rows:
        enriched = _cache_row_to_enriched(row)
        aid = enriched.get("artist_id")
        if not aid:
            continue
        # Merge: Supabase values overwrite local, but don't lose local-only fields
        existing = records.get(aid, {})
        existing.update({k: v for k, v in enriched.items() if v is not None})
        records[aid] = existing
        updated += 1

    _SCRAPER_DATA.mkdir(parents=True, exist_ok=True)
    with open(_ENRICHED_FILE, "w", encoding="utf-8") as f:
        for d in records.values():
            f.write(json.dumps(d, ensure_ascii=False) + "\n")

    booked = sum(1 for r in records.values() if r.get("lofi_booked"))
    print(f"  Written {len(records)} records ({booked} LOFI-booked) to {_ENRICHED_FILE}")
    print("Next: python run.py --seed")


def cmd_enrich(args) -> None:
    from data_aggregator import run_aggregation
    run_aggregation(verbose=True)
    cmd_export_excel(args)


def cmd_export_excel(args) -> None:
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    export_excel()


def cmd_seed(args) -> None:
    from dotenv import load_dotenv
    load_dotenv()

    from lofi_tinder.embedder import (
        build_index, compute_centroid, compute_feature_centroid,
        compute_dual_feature_centroids, embed_profiles, extract_feature_vector,
        save_centroid, save_feature_centroid, save_dual_feature_centroids, save_index,
    )
    from lofi_tinder.profile_builder import generate_profiles_batch

    enriched = _load_enriched()
    booked = [r for r in enriched.values() if r.get("lofi_booked")]
    print(f"Seeding from {len(booked)} LOFI-booked artists")

    artists = [_enriched_to_artist_input(r) for r in booked]
    profiles_map = generate_profiles_batch(artists, verbose=True)

    needs_embedding = [p for p in profiles_map.values() if not p.embedding]
    if needs_embedding:
        print(f"Embedding {len(needs_embedding)} profiles...")
        embed_profiles(needs_embedding)
        _flush_profile_embeddings(profiles_map)

    all_profiles = list(profiles_map.values())
    print(f"Building FAISS index from {len(all_profiles)} seed profiles...")
    index, id_map = build_index(all_profiles)
    save_index(index, id_map)

    seed_embeddings = [p.embedding for p in all_profiles if p.embedding]
    if seed_embeddings:
        centroid = compute_centroid(seed_embeddings)
        save_centroid(centroid)
        print(f"Text centroid saved from {len(seed_embeddings)} artists.")

    # Build structured feature centroid (15-dim, maps to Chartmetric filter params)
    feature_vecs = [extract_feature_vector(r) for r in booked]
    feature_centroid = compute_feature_centroid(feature_vecs)
    save_feature_centroid(feature_centroid)
    print(f"Feature centroid saved ({len(feature_vecs)} artists, {len(feature_centroid)} dims).")

    # Build dual centroids: core (established, top 50% by bookings) + emerging (bottom 50%)
    core_centroid, emerging_centroid = compute_dual_feature_centroids(booked)
    save_dual_feature_centroids(core_centroid, emerging_centroid)
    import numpy as np
    # Estimate the booking threshold used for the split
    totals = sorted(int((r.get("booking_stats") or {}).get("total") or 0) for r in booked)
    threshold = totals[len(totals) // 2] if totals else 0
    core_count     = sum(1 for t in totals if t >= threshold)
    emerging_count = sum(1 for t in totals if t <  threshold)
    print(f"Dual feature centroids saved: core={core_count} artists (≥{threshold} bookings), emerging={emerging_count} artists (<{threshold} bookings).")

    _preseed_swipes_and_mab(all_profiles, enriched)

    print("Seed complete. Run: python run.py --candidates")


def _preseed_swipes_and_mab(profiles: list, enriched: dict[str, dict]) -> None:
    import numpy as np
    from lofi_tinder.embedder import extract_feature_vector
    from lofi_tinder.mab import LinUCB
    from lofi_tinder.supabase_client import get_client as _supabase
    from lofi_tinder.schemas import SwipeRecord

    _SWIPES_FILE = _DATA_DIR / "swipes.jsonl"
    _DATA_DIR.mkdir(parents=True, exist_ok=True)

    existing = set()
    if _SWIPES_FILE.exists():
        for line in _SWIPES_FILE.read_text(encoding="utf-8").splitlines():
            if line.strip():
                try:
                    existing.add(json.loads(line)["artist_id"])
                except Exception:
                    pass

    sb = _supabase()
    if sb.available:
        print("Supabase connected — seed swipes will be pushed to DB.")
    else:
        print("Supabase not available — seed swipes saved to local file only.")

    mab = LinUCB.load()
    new_count = 0

    with open(_SWIPES_FILE, "a", encoding="utf-8") as f:
        for profile in profiles:
            if profile.artist_id in existing:
                continue
            ts = datetime.now(timezone.utc).isoformat()
            swipe = SwipeRecord(
                artist_id=profile.artist_id,
                name=profile.name,
                decision="yes",
                ts=ts,
                cosine_dist_at_swipe=profile.cosine_dist_to_centroid,
                linucb_score_at_swipe=0.0,
                profile_text=profile.profile_text,
            )
            f.write(swipe.model_dump_json() + "\n")
            if sb.available:
                try:
                    sb.save_swipe(
                        artist_id=profile.artist_id,
                        name=profile.name,
                        decision="yes",
                        ts=ts,
                        score=profile.cosine_dist_to_centroid,
                        profile_text=profile.profile_text,
                    )
                except Exception as exc:
                    print(f"  Supabase write failed for {profile.name}: {exc}")
            enr = enriched.get(profile.artist_id) or {}
            if enr:
                fvec = extract_feature_vector(enr)
                mab.update(fvec.astype("float64"), 1.0)
            new_count += 1

    mab.save()
    if sb.available:
        counts = sb.count_swipes()
        print(f"Supabase: {sum(counts.values())} total swipes ({counts.get('yes', 0)} YES).")
    print(f"Pre-seeded {new_count} YES swipes and MAB weights (14-dim features) from LOFI-booked artists.")


def cmd_candidates(args) -> None:
    from dotenv import load_dotenv
    load_dotenv()

    from lofi_tinder.embedder import embed_profiles, load_centroid
    from lofi_tinder.profile_builder import generate_profiles_batch

    enriched = _load_enriched()
    candidates_data = [r for r in enriched.values() if not r.get("lofi_booked")]
    print(f"Found {len(candidates_data)} candidate artists")

    artists = [_enriched_to_artist_input(r) for r in candidates_data]
    profiles_map = generate_profiles_batch(artists, verbose=True)

    needs_embedding = [p for p in profiles_map.values() if not p.embedding]
    if needs_embedding:
        print(f"Embedding {len(needs_embedding)} candidate profiles...")
        embed_profiles(needs_embedding)
        _flush_profile_embeddings(profiles_map)

    for artist in artists:
        _save_candidate(artist)

    print(f"Done. {len(artists)} candidates saved.")
    print("Run: streamlit run lofi_tinder/app.py")


def cmd_build_index(args) -> None:
    from lofi_tinder.embedder import build_index, embed_profiles, save_index
    from lofi_tinder.schemas import ArtistProfile

    all_profiles: dict[str, ArtistProfile] = {}
    if not _PROFILES_FILE.exists():
        print("No profiles found. Run --seed first.")
        return
    for line in _PROFILES_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            try:
                p = ArtistProfile(**json.loads(line))
                all_profiles[p.artist_id] = p
            except Exception:
                pass

    needs_embedding = [p for p in all_profiles.values() if not p.embedding]
    if needs_embedding:
        print(f"Embedding {len(needs_embedding)} profiles (sentence-transformers)...")
        embed_profiles(needs_embedding)
        _flush_profile_embeddings(all_profiles)
        print(f"Embeddings written.")

    embedded = [p for p in all_profiles.values() if p.embedding]
    print(f"Building FAISS index from {len(embedded)} profiles...")
    index, id_map = build_index(embedded)
    save_index(index, id_map)
    print("Index rebuilt.")


def cmd_retrain_mab(args) -> None:
    """Replay all existing swipes through the MAB using 14-dim feature vectors.
    Run this after upgrading from the old 384-dim MAB, or any time weights get stale."""
    from dotenv import load_dotenv
    load_dotenv()

    from lofi_tinder.embedder import extract_feature_vector
    from lofi_tinder.mab import LinUCB, reward_for_decision
    from lofi_tinder.ranker import load_swipes

    enriched = _load_enriched()
    swipes   = load_swipes()

    mab = LinUCB()   # fresh — shape is already _DIM=14
    count = 0
    skipped = 0
    for swipe in swipes:
        r = reward_for_decision(swipe.decision)
        if r is None:
            continue
        enr = enriched.get(swipe.artist_id) or {}
        if not enr:
            skipped += 1
            continue
        fvec = extract_feature_vector(enr)
        mab.update(fvec.astype("float64"), r)
        count += 1

    mab.save()
    print(f"MAB retrained: {count} swipes replayed, {skipped} skipped (no enriched data).")
    print(f"New 14-dim weights saved to {_DATA_DIR / 'mab_weights.npz'}")


def cmd_stats(args) -> None:
    from lofi_tinder.embedder import _CENTROID_FILE, _FEATURE_CENTROID_FILE, _INDEX_FILE
    from lofi_tinder.ranker import load_swipes

    enriched_count = 0
    if _ENRICHED_FILE.exists():
        enriched_count = sum(1 for l in _ENRICHED_FILE.read_text(encoding="utf-8").splitlines() if l.strip())

    profiles_count = 0
    if _PROFILES_FILE.exists():
        profiles_count = sum(1 for l in _PROFILES_FILE.read_text(encoding="utf-8").splitlines() if l.strip())

    candidates_count = 0
    if _CANDIDATES_FILE.exists():
        candidates_count = sum(1 for l in _CANDIDATES_FILE.read_text(encoding="utf-8").splitlines() if l.strip())

    swipes = load_swipes()
    yes_count = sum(1 for s in swipes if s.decision == "yes")
    no_count  = sum(1 for s in swipes if s.decision == "no")

    excels = list(_SCRAPER_DATA.glob("lofi_artist_data_*.xlsx"))
    latest_excel = max(excels, key=lambda p: p.stat().st_mtime).name if excels else "none"

    print(f"\nLOFI Tinder - status")
    print(f"  Enriched records:   {enriched_count}")
    print(f"  Profiles cached:    {profiles_count}")
    print(f"  Candidates queued:  {candidates_count}")
    print(f"  FAISS index:        {'OK' if _INDEX_FILE.exists() else 'MISSING - run --build-index'}")
    print(f"  Text centroid:      {'OK' if _CENTROID_FILE.exists() else 'MISSING - run --seed'}")
    print(f"  Feature centroid:   {'OK' if _FEATURE_CENTROID_FILE.exists() else 'MISSING - run --seed'}")
    print(f"  Swipes:             {len(swipes)} total ({yes_count} YES, {no_count} NO)")
    print(f"  Latest Excel:       {latest_excel}")


def main() -> None:
    parser = argparse.ArgumentParser(description="LOFI Tinder entrypoint")
    parser.add_argument("--pull",         action="store_true", help="Sync Supabase tinder.artist_cache -> local artist_enriched.jsonl")
    parser.add_argument("--enrich",       action="store_true", help="Aggregate local scraper data into artist_enriched.jsonl")
    parser.add_argument("--export-excel", action="store_true", help="Export scraper_data to Excel")
    parser.add_argument("--seed",         action="store_true", help="Generate profiles for LOFI-booked artists + rebuild centroids")
    parser.add_argument("--candidates",   action="store_true", help="Generate profiles for candidates")
    parser.add_argument("--build-index",  action="store_true", help="Rebuild FAISS index")
    parser.add_argument("--retrain-mab",  action="store_true", help="Replay all swipes through 14-dim MAB")
    parser.add_argument("--stats",        action="store_true", help="Show status summary")
    args = parser.parse_args()

    if args.pull:
        cmd_pull(args)
    elif args.enrich:
        cmd_enrich(args)
    elif getattr(args, "export_excel"):
        cmd_export_excel(args)
    elif args.seed:
        cmd_seed(args)
    elif args.candidates:
        cmd_candidates(args)
    elif getattr(args, "build_index"):
        cmd_build_index(args)
    elif getattr(args, "retrain_mab"):
        cmd_retrain_mab(args)
    elif args.stats:
        cmd_stats(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
