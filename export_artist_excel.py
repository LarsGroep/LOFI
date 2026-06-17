"""Export all scraped data for a single artist to a multi-sheet Excel file."""
import os
import json
import argparse
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

from supabase import create_client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="F3F4F6")


def to_cell(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def jsonb(val):
    if isinstance(val, str):
        try:
            return json.loads(val)
        except Exception:
            return []
    return val or []


def flat(d, prefix="", skip=("id", "artist_id", "updated_at", "created_at")):
    return {
        (prefix + k if prefix else k): to_cell(v)
        for k, v in d.items()
        if k not in skip
    }


def flatten_row(row):
    return {k: to_cell(v) for k, v in row.items()}


def style_sheet(ws, rows, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, str(h))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, to_cell(val))
            if ri % 2 == 0:
                c.fill = ALT_FILL
    for ci in range(1, len(headers) + 1):
        vals = [str(ws.cell(r, ci).value or "") for r in range(1, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max((len(v) for v in vals), default=8) + 2, 70
        )
    ws.freeze_panes = "A2"


def jsonb_sheet(ws, data, sort_key=None, sort_rev=True):
    if not data or not isinstance(data[0], dict):
        return 0
    data = [flatten_row(r) for r in data]
    if sort_key:
        data.sort(key=lambda x: x.get(sort_key) or 0, reverse=sort_rev)
    cols = list(data[0].keys())
    style_sheet(ws, [[r.get(c, "") for c in cols] for r in data], cols)
    return len(data)


def kv_sheet(ws, d):
    style_sheet(ws, [[k, to_cell(v)] for k, v in d.items() if v not in (None, "", [])], ["Metric", "Value"])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 90


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--artist-id", required=True)
    parser.add_argument("--out", default="artist_export.xlsx")
    args = parser.parse_args()

    sb  = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])
    aid = args.artist_id

    def get(table, id_col="artist_id"):
        return sb.schema("tinder").table(table).select("*").eq(id_col, aid).execute().data or []

    artist     = get("artists", "id")[0]
    cm         = get("artist_chartmetric")[0]
    ext        = get("artist_cm_extended")[0]
    ra         = get("artist_ra")[0]
    lfm        = get("artist_lastfm")[0]
    pf         = get("artist_partyflock")[0]
    ra_events  = get("ra_events")
    tracks     = get("artist_cm_tracks")
    val_events = get("validation_events")

    name = artist.get("name", aid)
    wb   = openpyxl.Workbook()

    # ── 1. Profile (transposed) ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Profile"
    profile = {}
    profile.update(flat(artist, skip=("id", "updated_at", "created_at")))
    profile.update(flat(cm,  "cm_"))
    profile.update(flat(ext, "ext_"))
    profile.update(flat(ra,  "ra_"))
    profile.update(flat(lfm, "lfm_"))
    profile.update(flat(pf,  "pf_"))
    kv_sheet(ws, profile)
    print(f"  Profile: {len(profile)} fields")

    # ── 2. RA Events (with full lineup) ──────────────────────────────────────
    ws2   = wb.create_sheet("RA Events")
    cols  = ["date", "title", "venue", "city", "country", "venue_capacity", "lineup_size", "lineup", "event_url"]
    rows  = []
    for e in sorted(ra_events, key=lambda x: x.get("date", ""), reverse=True):
        lineup = jsonb(e.get("lineup"))
        rows.append([
            e.get("date"), e.get("title"), e.get("venue"), e.get("city"),
            e.get("country"), e.get("venue_capacity"), e.get("lineup_size"),
            ", ".join(lineup) if lineup else "",
            e.get("event_url"),
        ])
    style_sheet(ws2, rows, cols)
    print(f"  RA Events: {len(rows)} rows")

    # ── 3. Partyflock Events ──────────────────────────────────────────────────
    ws3      = wb.create_sheet("Partyflock Events")
    pf_evs   = jsonb(pf.get("events"))
    pf_cols  = ["start_date", "event_name", "venue", "city", "country", "lineup_size", "event_url"]
    pf_rows  = [
        [to_cell(e.get(c)) for c in pf_cols]
        for e in sorted(pf_evs, key=lambda x: x.get("start_date", ""), reverse=True)
        if isinstance(e, dict)
    ]
    style_sheet(ws3, pf_rows, pf_cols)
    print(f"  Partyflock Events: {len(pf_rows)} rows")

    # ── 4. Last.fm ────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Last.fm")
    lfm_d = {
        "listeners":       lfm.get("lfm_listeners"),
        "playcount":       lfm.get("lfm_playcount"),
        "tags":            ", ".join(lfm.get("tags") or []),
        "similar_artists": ", ".join(lfm.get("similar_artists") or []),
    }
    kv_sheet(ws4, lfm_d)
    print("  Last.fm: summary")

    # ── 5. Tracks ────────────────────────────────────────────────────────────
    ws5    = wb.create_sheet("Tracks")
    tr_cols = ["track_name", "isrc", "release_date", "spotify_streams",
               "spotify_popularity", "peak_spotify_chart", "peak_beatport_chart", "playlist_count"]
    tr_rows = [
        [to_cell(t.get(c)) for c in tr_cols]
        for t in sorted(tracks, key=lambda x: x.get("spotify_streams") or 0, reverse=True)
    ]
    style_sheet(ws5, tr_rows, tr_cols)
    print(f"  Tracks: {len(tr_rows)} rows")

    # ── 6. Albums ────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Albums")
    n   = jsonb_sheet(ws6, jsonb(ext.get("albums")))
    print(f"  Albums: {n} rows")

    # ── 7. Venues ────────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Venues")
    n   = jsonb_sheet(ws7, jsonb(ext.get("venues")))
    print(f"  Venues: {n} rows")

    # ── 8. Career History ─────────────────────────────────────────────────────
    ws8 = wb.create_sheet("Career History")
    n   = jsonb_sheet(ws8, jsonb(ext.get("career_history")), "timestp", True)
    print(f"  Career History: {n} rows")

    # ── 9. CM Stats (weekly/monthly diffs) ────────────────────────────────────
    ws9     = wb.create_sheet("CM Stats")
    cm_stats = ext.get("cm_stats")
    if isinstance(cm_stats, str):
        try:
            cm_stats = json.loads(cm_stats)
        except Exception:
            cm_stats = {}
    if cm_stats and isinstance(cm_stats, dict):
        kv_sheet(ws9, cm_stats)
        print(f"  CM Stats: {len(cm_stats)} metrics")

    # ── 10. Related Artists ───────────────────────────────────────────────────
    ws10 = wb.create_sheet("Related Artists")
    n    = jsonb_sheet(ws10, jsonb(ext.get("related_artists")), "cm_artist_score")
    print(f"  Related Artists: {n} rows")

    # ── 11. CM Milestones ────────────────────────────────────────────────────
    ws11 = wb.create_sheet("CM Milestones")
    n    = jsonb_sheet(ws11, jsonb(ext.get("milestones")))
    print(f"  CM Milestones: {n} rows")

    # ── 12. Noteworthy Insights ───────────────────────────────────────────────
    ws12 = wb.create_sheet("Insights")
    n    = jsonb_sheet(ws12, jsonb(ext.get("noteworthy_insights")))
    print(f"  Insights: {n} rows")

    # ── 13. News ──────────────────────────────────────────────────────────────
    ws13 = wb.create_sheet("News")
    n    = jsonb_sheet(ws13, jsonb(ext.get("news")))
    print(f"  News: {n} rows")

    # ── 14. External Events (Songkick/SeatGeek) ───────────────────────────────
    ws14 = wb.create_sheet("External Events")
    n    = jsonb_sheet(ws14, jsonb(ext.get("events_external")))
    print(f"  External Events: {n} rows")

    # ── 15. Platform URLs ─────────────────────────────────────────────────────
    ws15 = wb.create_sheet("Platform URLs")
    urls = ext.get("urls")
    if isinstance(urls, str):
        try:
            urls = json.loads(urls)
        except Exception:
            urls = {}
    if urls and isinstance(urls, dict):
        kv_sheet(ws15, {k: v for k, v in urls.items() if v})
        print(f"  Platform URLs: {len([v for v in urls.values() if v])} rows")

    # ── 16. Validation Events ─────────────────────────────────────────────────
    ws16   = wb.create_sheet("Validation Events")
    ve_cols = ["event_type", "event_date", "source", "confirmed"]
    style_sheet(ws16, [[to_cell(e.get(c)) for c in ve_cols] for e in val_events], ve_cols)
    print(f"  Validation Events: {len(val_events)} rows")

    wb.save(args.out)
    print(f"\nSaved: {args.out}  ({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
