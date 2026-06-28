import os, sys, json
from pathlib import Path
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent / 'scrapers/live_detection'))
from shared.db import get_client

import pandas as pd

sb = get_client()

# ── 1. Beatport chart entries ────────────────────────────────────────────
bp_raw = sb.schema('tinder').table('beatport_chart_entries').select('*').order('genre').order('chart_position').execute()
df_bp = pd.DataFrame(bp_raw.data)
df_bp['scraped_at'] = pd.to_datetime(df_bp['scraped_at']).dt.strftime('%Y-%m-%d')
df_bp['scouted'] = df_bp['artist_id'].notna()

BEATPORT_GENRES = {
    'tech-house':               'Tech House',
    'house':                    'House',
    'techno-peak-time-driving': 'Techno (Peak Time)',
    'melodic-house-techno':     'Melodic House & Techno',
    'minimal-deep-tech':        'Minimal / Deep Tech',
    'afro-house':               'Afro House',
    'organic-house-downtempo':  'Organic House',
    'progressive-house':        'Progressive House',
}
df_bp['genre_label'] = df_bp['genre'].map(BEATPORT_GENRES).fillna(df_bp['genre'])
df_bp_out = df_bp[['chart_position', 'artist_name', 'track_name', 'genre_label', 'scouted', 'scraped_at']].copy()
df_bp_out.columns = ['Position', 'Artist', 'Track', 'Genre', 'Scouted Artist?', 'Scraped']

# ── 2. Traxsource chart entries ──────────────────────────────────────────
tx_raw = sb.schema('tinder').table('traxsource_chart_entries').select('*').order('genre').order('chart_position').execute()
df_tx = pd.DataFrame(tx_raw.data)
df_tx['scraped_at'] = pd.to_datetime(df_tx['scraped_at']).dt.strftime('%Y-%m-%d')
df_tx['scouted'] = df_tx['artist_id'].notna()

TRAXSOURCE_GENRES = {
    'tech-house':                'Tech House',
    'house':                     'House',
    'deep-house':                'Deep House',
    'techno':                    'Techno',
    'minimal-deep-tech':         'Minimal / Deep Tech',
    'melodic-progressive-house': 'Melodic / Progressive House',
    'afro-house':                'Afro House',
    'nu-disco-indie-dance':      'Nu Disco / Indie Dance',
}
df_tx['genre_label'] = df_tx['genre'].map(TRAXSOURCE_GENRES).fillna(df_tx['genre'])
df_tx_out = df_tx[['chart_position', 'artist_name', 'track_name', 'genre_label', 'scouted', 'scraped_at']].copy()
df_tx_out.columns = ['Position', 'Artist', 'Track', 'Genre', 'Scouted Artist?', 'Scraped']

# ── 3. Milestones ────────────────────────────────────────────────────────
ms_raw = sb.schema('tinder').table('validation_events').select(
    'artist_id, event_type, event_date, source, details'
).in_('event_type', [
    'first_beatport_chart', 'first_beatport_top_10', 'first_beatport_number_1',
    'first_traxsource_chart', 'first_traxsource_top_10', 'first_traxsource_number_1',
]).order('event_date', desc=True).execute()

artists_raw = sb.schema('tinder').table('artist_chartmetric_flat').select('artist_id, artist_name').execute()
artist_name_map = {r['artist_id']: r['artist_name'] for r in artists_raw.data}

rows_ms = []
for r in ms_raw.data:
    details = r.get('details') or {}
    if isinstance(details, str):
        try:
            details = json.loads(details)
        except Exception:
            details = {}
    rows_ms.append({
        'Artist':    artist_name_map.get(r['artist_id'], r['artist_id']),
        'Milestone': r['event_type'].replace('first_', '').replace('_', ' ').title(),
        'Date':      r['event_date'],
        'Source':    r['source'],
        'Genre':     details.get('genre', ''),
        'Position':  details.get('position', ''),
        'Track':     details.get('track', ''),
    })

df_ms = pd.DataFrame(rows_ms).sort_values(['Artist', 'Date'])

# ── 4. Scouted artists summary ───────────────────────────────────────────
bp_scouted = df_bp[df_bp['scouted']].copy()
tx_scouted = df_tx[df_tx['scouted']].copy()

summary = defaultdict(lambda: {
    'bp_genres': set(), 'bp_best_pos': 999,
    'tx_genres': set(), 'tx_best_pos': 999,
})
for _, row in bp_scouted.iterrows():
    s = summary[row['artist_name']]
    s['bp_genres'].add(row['genre_label'])
    if row['chart_position']:
        s['bp_best_pos'] = min(s['bp_best_pos'], row['chart_position'])

for _, row in tx_scouted.iterrows():
    s = summary[row['artist_name']]
    s['tx_genres'].add(row['genre_label'])
    if row['chart_position']:
        s['tx_best_pos'] = min(s['tx_best_pos'], row['chart_position'])

summary_rows = []
for name, d in sorted(summary.items()):
    summary_rows.append({
        'Artist':                   name,
        'Beatport — Genres':        ', '.join(sorted(d['bp_genres'])) or '—',
        'Beatport — Best Position': d['bp_best_pos'] if d['bp_best_pos'] < 999 else '—',
        'Traxsource — Genres':      ', '.join(sorted(d['tx_genres'])) or '—',
        'Traxsource — Best Pos':    d['tx_best_pos'] if d['tx_best_pos'] < 999 else '—',
    })

df_summary = pd.DataFrame(summary_rows)

# ── 5. Write Excel ───────────────────────────────────────────────────────
today = datetime.now().strftime('%Y-%m-%d')
outpath = Path(__file__).parent / f'chart_scraper_output_{today}.xlsx'

with pd.ExcelWriter(outpath, engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name='Scouted Artists on Charts', index=False)
    df_ms.to_excel(writer, sheet_name='Milestones', index=False)
    df_bp_out.to_excel(writer, sheet_name='Beatport Top 100', index=False)
    df_tx_out.to_excel(writer, sheet_name='Traxsource Top 100', index=False)

    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    highlight = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bold_font = Font(bold=True)

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)
        for cell in ws[1]:
            cell.font = bold_font
        if sheet_name in ('Beatport Top 100', 'Traxsource Top 100'):
            scouted_col = None
            for cell in ws[1]:
                if cell.value == 'Scouted Artist?':
                    scouted_col = cell.column
                    break
            if scouted_col:
                for row in ws.iter_rows(min_row=2):
                    if row[scouted_col - 1].value is True:
                        for cell in row:
                            cell.fill = highlight

print(f'Saved: {outpath}')
print(f'  Scouted Artists on Charts: {len(df_summary)} artists')
print(f'  Milestones sheet: {len(df_ms)} rows')
print(f'  Beatport Top 100: {len(df_bp_out)} rows')
print(f'  Traxsource Top 100: {len(df_tx_out)} rows')
